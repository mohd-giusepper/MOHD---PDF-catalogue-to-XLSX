import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from pdf2xlsx import config
from pdf2xlsx.core import auto_convert, page_cache
from pdf2xlsx.io import xlsx_writer
from pdf2xlsx.models import ProductRow, RunReport, TriageResult


class SamplingTest(unittest.TestCase):
    def test_sample_count_buckets(self) -> None:
        self.assertEqual(page_cache.compute_sample_count(40), 8)
        self.assertEqual(page_cache.compute_sample_count(150), 8)
        self.assertEqual(page_cache.compute_sample_count(350), 8)
        self.assertEqual(page_cache.compute_sample_count(900), 9)

    def test_strong_hit_requires_code_or_cooccurrence(self) -> None:
        page = page_cache.CachedPage(
            page_number=1,
            text="EUR 100",
            normalized_text="EUR 100",
            lines=["EUR 100"],
            words=[],
            text_len=10,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
            signal_score=config.CACHE_EARLY_STOP_SCORE + 1.0,
            mixed_code_count=0,
            cooccurrence_count=0,
        )
        self.assertFalse(page_cache._is_strong_hit(page, has_stopword=False))
        page.mixed_code_count = 1
        self.assertTrue(page_cache._is_strong_hit(page, has_stopword=False))

    def test_retry_on_zero_cooccurrence_resamples(self) -> None:
        dummy_page = page_cache.CachedPage(
            page_number=1,
            text="EUR 100",
            normalized_text="EUR 100",
            lines=["EUR 100"],
            words=[],
            text_len=10,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
            cooccurrence_count=0,
            mixed_code_count=0,
        )
        retry_page = page_cache.CachedPage(
            page_number=5,
            text="AB12 EUR 100",
            normalized_text="AB12 EUR 100",
            lines=["AB12 EUR 100"],
            words=[],
            text_len=12,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
            cooccurrence_count=1,
            mixed_code_count=1,
        )
        meta_first = {
            "sample_count": 10,
            "pages_sampled": [1],
            "num_pages": 10,
            "cooccurrence_pages_selected": 0,
            "cooccurrence_min_required": 1,
        }
        meta_second = {
            "sample_count": 20,
            "pages_sampled": [5],
            "num_pages": 10,
            "cooccurrence_pages_selected": 1,
            "cooccurrence_min_required": 1,
        }

        original_sweep = config.CACHE_ENABLE_SWEEP_AFTER_RETRY
        config.CACHE_ENABLE_SWEEP_AFTER_RETRY = False
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                with mock.patch(
                    "pdf2xlsx.core.page_cache.build_signal_cache",
                    side_effect=[
                        ([dummy_page], [], meta_first),
                        ([retry_page], [], meta_second),
                    ],
                ) as build_mock:
                    with mock.patch(
                        "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                        return_value={
                            "parser": "table_based",
                            "ok": True,
                            "reason": "",
                            "metrics": {
                                "eval_score": 1.0,
                                "eval_rows": 5,
                                "eval_pages_used": 1,
                                "eval_time_ms": 1,
                                "key_fields_rate": 0.9,
                                "review_rate": 0.1,
                            },
                            "score": 1.0,
                        },
                    ):
                        result = auto_convert.run_auto_for_pdf(
                            pdf_path="dummy.pdf",
                            output_dir=temp_dir,
                            fast_eval_only=True,
                        )
                self.assertEqual(build_mock.call_count, 2)
                self.assertTrue(result.sampling_retry_triggered)
                self.assertEqual(result.sampling_retry_reason, "cooccurrence_shortfall")
                self.assertTrue(
                    build_mock.call_args_list[1].kwargs.get(
                        "exclude_zero_cooccurrence_pages"
                    )
                )
        finally:
            config.CACHE_ENABLE_SWEEP_AFTER_RETRY = original_sweep

    def test_retry_on_too_few_rows_logs(self) -> None:
        dummy_page = page_cache.CachedPage(
            page_number=1,
            text="",
            normalized_text="",
            lines=[],
            words=[],
            text_len=0,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
        )
        cached_pages = [dummy_page]
        meta_first = {"sample_count": 10, "pages_sampled": [1], "num_pages": 10}
        meta_second = {"sample_count": 20, "pages_sampled": [1, 2], "num_pages": 10}

        original_sweep = config.CACHE_ENABLE_SWEEP_AFTER_RETRY
        config.CACHE_ENABLE_SWEEP_AFTER_RETRY = False
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                with mock.patch(
                    "pdf2xlsx.core.page_cache.build_signal_cache",
                    side_effect=[(cached_pages, [], meta_first), (cached_pages, [], meta_second)],
                ):
                    with mock.patch(
                        "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                        return_value={
                            "parser": "table_based",
                            "ok": False,
                            "reason": "too_few_rows",
                            "metrics": {"eval_score": 0.0},
                        },
                    ):
                        with self.assertLogs(
                            "pdf2xlsx.core.auto_convert", level="INFO"
                        ) as logs:
                            auto_convert.run_auto_for_pdf(
                                pdf_path="dummy.pdf",
                                output_dir=temp_dir,
                                fast_eval_only=True,
                            )
            self.assertTrue(
                any("RETRY_SCAN" in message for message in logs.output),
                "Missing RETRY_SCAN log entry",
            )
        finally:
            config.CACHE_ENABLE_SWEEP_AFTER_RETRY = original_sweep

    def test_retry_on_all_parsers_early_discard_with_cached_pages(self) -> None:
        dummy_page = page_cache.CachedPage(
            page_number=1,
            text="EUR 100",
            normalized_text="EUR 100",
            lines=["EUR 100"],
            words=[],
            text_len=10,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
            cooccurrence_count=0,
            mixed_code_count=0,
        )
        retry_page = page_cache.CachedPage(
            page_number=4,
            text="AB12 EUR 100",
            normalized_text="AB12 EUR 100",
            lines=["AB12 EUR 100"],
            words=[],
            text_len=12,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
            cooccurrence_count=1,
            mixed_code_count=1,
        )
        meta_retry = {
            "sample_count": 12,
            "pages_sampled": [4],
            "num_pages": 12,
            "cooccurrence_pages_selected": 1,
            "cooccurrence_min_required": 1,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch(
                "pdf2xlsx.core.page_cache.build_signal_cache",
                return_value=([retry_page], [], meta_retry),
            ) as build_mock:
                with mock.patch(
                    "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                    return_value={
                        "parser": "table_based",
                        "ok": False,
                        "reason": "early_discard_no_rows",
                        "metrics": {"eval_score": 0.0, "eval_rows": 0},
                    },
                ):
                    result = auto_convert.run_auto_for_pdf(
                        pdf_path="dummy.pdf",
                        output_dir=temp_dir,
                        fast_eval_only=True,
                        cached_pages=[dummy_page],
                    )
        self.assertEqual(build_mock.call_count, 1)
        self.assertTrue(result.sampling_retry_triggered)
        self.assertEqual(
            result.sampling_retry_reason, "all_parsers_early_discard_no_rows"
        )
        self.assertEqual(result.cached_pages_source, "resampled")

    def test_eval_pages_increase_on_weak_signals(self) -> None:
        cached_pages = []
        for idx in range(config.AUTO_EVAL_MIN_PAGES_WEAK):
            cached_pages.append(
                page_cache.CachedPage(
                    page_number=idx + 1,
                    text="foo",
                    normalized_text="foo",
                    lines=["foo"],
                    words=[],
                    text_len=3,
                    images_count=0,
                    needs_ocr=False,
                    ocr_used=False,
                    table_hint=False,
                    cooccurrence_count=0,
                    mixed_code_count=0,
                )
            )
        result = auto_convert.evaluate_parser_fast(
            cached_pages=cached_pages,
            parser_name="code_price_based",
            source_file="dummy.pdf",
            currency="EUR",
            currency_only=None,
            triage_result=None,
        )
        self.assertEqual(result.get("reason"), "early_discard_no_rows")
        self.assertGreaterEqual(
            result.get("metrics", {}).get("eval_pages_used", 0),
            config.AUTO_EVAL_MIN_PAGES_WEAK,
        )

    def test_retry_min_top_k_with_soft_toc_pages(self) -> None:
        pages = []
        total_pages = config.CACHE_RETRY_MIN_TOP_K + 2
        for idx in range(total_pages):
            pages.append(
                page_cache.CachedPage(
                    page_number=idx + 1,
                    text="toc page",
                    normalized_text="toc page",
                    lines=["100 2026 EUR"],
                    words=[],
                    text_len=120,
                    images_count=0,
                    needs_ocr=False,
                    ocr_used=False,
                    table_hint=False,
                    toc_like=True,
                    toc_hard=False,
                    row_candidate_count=2 if idx < config.CACHE_RETRY_MIN_TOP_K else 0,
                    cooccurrence_near_count=1 if idx < config.CACHE_RETRY_MIN_TOP_K else 0,
                    mixed_code_count=1 if idx < config.CACHE_RETRY_MIN_TOP_K else 0,
                    signal_score=5.0 + idx,
                )
            )
        selected, meta = page_cache._select_top_k_pages(
            pages=pages,
            top_k_count=1,
            min_cooccurrence_pages=0,
            exclude_zero_cooccurrence_pages=False,
            exclude_toc_pages=True,
            min_top_k_count=config.CACHE_RETRY_MIN_TOP_K,
        )
        self.assertGreaterEqual(len(selected), config.CACHE_RETRY_MIN_TOP_K)
        self.assertTrue(any(page.row_candidate_count > 0 for page in selected))
        self.assertFalse(meta.get("top_k_collapse_reason"))

    def test_diagnostic_xlsx_written_on_failure(self) -> None:
        dummy_page = page_cache.CachedPage(
            page_number=1,
            text="",
            normalized_text="",
            lines=[],
            words=[],
            text_len=0,
            images_count=0,
            needs_ocr=False,
            ocr_used=False,
            table_hint=False,
        )
        cached_pages = [dummy_page]
        meta = {
            "sample_count": 10,
            "pages_sampled": [1],
            "num_pages": 10,
            "top_k_pages": [1],
            "top_k_scores": [1.0],
        }

        original_sweep = config.CACHE_ENABLE_SWEEP_AFTER_RETRY
        config.CACHE_ENABLE_SWEEP_AFTER_RETRY = False
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                with mock.patch(
                    "pdf2xlsx.core.page_cache.build_signal_cache",
                    return_value=(cached_pages, [], meta),
                ):
                    with mock.patch(
                        "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                        return_value={
                            "parser": "table_based",
                            "ok": False,
                            "reason": "too_few_rows",
                            "metrics": {"eval_score": 0.0},
                        },
                    ):
                        auto_convert.run_auto_for_pdf(
                            pdf_path="dummy.pdf",
                            output_dir=temp_dir,
                            fast_eval_only=True,
                        )
                diagnostic_path = os.path.join(temp_dir, "dummy.diagnostic.xlsx")
                self.assertTrue(os.path.exists(diagnostic_path))
        finally:
            config.CACHE_ENABLE_SWEEP_AFTER_RETRY = original_sweep

    def test_diagnostic_summary_populated_with_cached_pages(self) -> None:
        cached_pages = [
            page_cache.CachedPage(
                page_number=1,
                text="foo",
                normalized_text="foo",
                lines=["foo"],
                words=[],
                text_len=3,
                images_count=0,
                needs_ocr=False,
                ocr_used=False,
                table_hint=False,
                signal_score=1.2,
            ),
            page_cache.CachedPage(
                page_number=2,
                text="bar",
                normalized_text="bar",
                lines=["bar"],
                words=[],
                text_len=3,
                images_count=0,
                needs_ocr=False,
                ocr_used=False,
                table_hint=False,
                signal_score=0.8,
            ),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch(
                "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                return_value={
                    "parser": "table_based",
                    "ok": False,
                    "reason": "too_few_rows",
                    "metrics": {"eval_score": 0.0, "eval_rows": 1},
                },
            ):
                auto_convert.run_auto_for_pdf(
                    pdf_path="dummy.pdf",
                    output_dir=temp_dir,
                    fast_eval_only=True,
                    cached_pages=cached_pages,
                )
            diagnostic_path = os.path.join(temp_dir, "dummy.diagnostic.xlsx")
            workbook = load_workbook(diagnostic_path)
            summary_sheet = workbook["SUMMARY"]
            summary = {row[0].value: row[1].value for row in summary_sheet.iter_rows()}
            self.assertGreater(int(summary.get("num_pages", 0)), 0)
            self.assertGreater(int(summary.get("sample_count", 0)), 0)
            self.assertTrue(summary.get("pages_sampled"))

    def test_too_few_rows_low_eval_pages_passes(self) -> None:
        metrics = {
            "rows_exported": 1,
            "unique_code_ratio": 1.0,
            "key_fields_rate": 1.0,
            "review_rate": 0.0,
            "duplicate_conflicts_rate": 0.0,
            "eval_pages_used": config.AUTO_EVAL_MIN_PAGES_WEAK,
        }
        ok, reason, _ = auto_convert.evaluate_metrics(metrics)
        self.assertTrue(ok)
        self.assertEqual(reason, "passed")

    def test_review_only_output_when_eval_rows_present(self) -> None:
        cached_pages = [
            page_cache.CachedPage(
                page_number=1,
                text="AB12 EUR 100",
                normalized_text="AB12 EUR 100",
                lines=["AB12 EUR 100"],
                words=[],
                text_len=12,
                images_count=0,
                needs_ocr=False,
                ocr_used=False,
                table_hint=False,
            )
        ]
        triage_result = TriageResult(suggested_profile="code_price_based", decision="OK")

        def fake_run_pipeline(*args, **kwargs):
            output_xlsx = kwargs.get("output_xlsx")
            rows = [
                ProductRow(
                    product_name_en="Lamp",
                    art_no="AB12",
                    price_eur=100.0,
                    exported=False,
                    needs_review=True,
                )
            ]
            if output_xlsx:
                xlsx_writer.write_xlsx(rows, output_xlsx)
            return RunReport(
                rows=rows,
                rows_exported=0,
                rows_needs_review=1,
                skipped_missing_target_price=0,
                duplicate_art_no_count=0,
                duplicate_conflicts_count=0,
                bad_art_no_count=0,
                corrected_art_no_count=0,
                suspicious_numeric_art_no_seen=False,
                examples_bad_art_no=[],
            )

        def fake_eval_fast(*args, **kwargs):
            parser_name = kwargs.get("parser_name")
            if parser_name == "code_price_based":
                return {
                    "parser": parser_name,
                    "ok": False,
                    "reason": "too_few_rows",
                    "metrics": {"eval_rows": 4, "eval_pages_used": 4, "eval_score": 0.4},
                    "score": 0.4,
                }
            return {
                "parser": parser_name,
                "ok": False,
                "reason": "early_discard_no_rows",
                "metrics": {"eval_rows": 0, "eval_pages_used": 2, "eval_score": 0.0},
                "score": 0.0,
            }

        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch(
                "pdf2xlsx.core.auto_convert.evaluate_parser_fast",
                side_effect=fake_eval_fast,
            ):
                with mock.patch(
                    "pdf2xlsx.core.pipeline.run_pipeline",
                    side_effect=fake_run_pipeline,
                ):
                    result = auto_convert.run_auto_for_pdf(
                        pdf_path="dummy.pdf",
                        output_dir=temp_dir,
                        fast_eval_only=False,
                        cached_pages=cached_pages,
                        triage_result=triage_result,
                    )
            final_xlsx = os.path.join(temp_dir, "dummy.xlsx")
            self.assertTrue(os.path.exists(final_xlsx))
            self.assertIn("REVIEW_ONLY", result.final_status)


if __name__ == "__main__":
    unittest.main()
