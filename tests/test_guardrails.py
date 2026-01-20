import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from pdf2xlsx.core import pipeline
from pdf2xlsx.io import xlsx_writer
from pdf2xlsx.models import ProductRow
from pdf2xlsx.utils import text as text_utils


class GuardrailTests(unittest.TestCase):
    def test_page_gate_skips_legal_text(self) -> None:
        lines = [
            "Terms and conditions apply to this catalogue.",
            "Articolo 12 del codice civile e decreto legislativo 123.",
            "Warranty and liability; jurisdiction: EU.",
        ]
        signals = pipeline.compute_page_signals(lines)
        self.assertTrue(pipeline.is_text_like_page(signals, "stelton_2025"))

    def test_numeric_art_no_with_price_passes(self) -> None:
        row = ProductRow(
            product_name_en="Sample chair",
            art_no="1381",
            price_eur=120.0,
        )
        raw_text = "1381 Sample chair EUR 120"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertTrue(ok)
        self.assertEqual(reason, "")

    def test_ambiguous_numeric_token_flags_review(self) -> None:
        row = ProductRow(
            product_name_en="Chair",
            art_no="12345",
            price_eur=12345.0,
        )
        raw_text = "12345 Chair"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        self.assertTrue(token_info.get("ambiguous_numeric"))
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertFalse(ok)
        self.assertEqual(reason, "ambiguous_price_vs_artno")

    def test_dimension_detection(self) -> None:
        raw_text = "AB12 Table 120 x 60 cm EUR 200"
        token_info = text_utils.resolve_row_fields(raw_text)
        dims = token_info.get("dimension_candidates") or []
        self.assertTrue(dims)
        self.assertIn("120 x 60 cm", dims[0])

    def test_legallike_row_goes_to_review(self) -> None:
        row = ProductRow(
            product_name_en="Legal note",
            art_no="AB1234",
            price_eur=120.0,
        )
        raw_text = "Ai sensi dell'articolo 12 della Direttiva 2003/35/EC EUR 120"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertFalse(ok)
        self.assertEqual(reason, "hard_legal")

    def test_dim_only_row_goes_to_review(self) -> None:
        row = ProductRow(
            product_name_en="prof.",
            art_no="1150",
            price_eur=1400.0,
        )
        raw_text = "1150 prof. prof. EUR 1400"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertFalse(ok)
        self.assertEqual(reason, "dim_only_row")

    def test_numeric_hyphen_art_no_passes(self) -> None:
        row = ProductRow(
            product_name_en="Lamp",
            art_no="010-1",
            price_eur=120.0,
        )
        raw_text = "010-1 Lamp EUR 120"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertTrue(ok)
        self.assertEqual(reason, "")

    def test_valid_row_passes_hard_validation(self) -> None:
        row = ProductRow(
            product_name_en="Lounge chair",
            art_no="AB1234",
            price_eur=350.0,
        )
        raw_text = "AB1234 Lounge chair EUR 350"
        line_info = text_utils.analyze_line(raw_text)
        token_info = text_utils.resolve_row_fields(raw_text)
        ok, reason = pipeline.hard_validate_row(
            row,
            raw_text,
            line_info,
            token_info=token_info,
        )
        self.assertTrue(ok)
        self.assertEqual(reason, "")

    def test_writer_exports_only_exported_rows(self) -> None:
        rows = [
            ProductRow(
                product_name_en="Chair",
                art_no="AB1234",
                price_eur=120.0,
                exported=True,
            ),
            ProductRow(
                product_name_en="Legal text",
                art_no="",
                price_eur=None,
                exported=False,
                needs_review=True,
            ),
        ]
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            output_path = Path(temp_file.name)
        try:
            xlsx_writer.write_xlsx(rows, str(output_path))
            workbook = load_workbook(str(output_path))
            products = workbook["PRODUCTS"]
            self.assertEqual(products.max_row, 2)
        finally:
            if output_path.exists():
                output_path.unlink()


if __name__ == "__main__":
    unittest.main()
