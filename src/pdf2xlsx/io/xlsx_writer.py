from typing import Iterable, List

from openpyxl import Workbook

from pdf2xlsx.models import ProductRow
from pdf2xlsx.utils import text as text_utils


HEADERS: List[str] = [
    "source_file",
    "page",
    "section",
    "product_name_en",
    "product_name_raw",
    "variant",
    "designer",
    "art_no",
    "colli",
    "size_raw",
    "width_cm",
    "height_cm",
    "length_cm",
    "price_dkk",
    "price_sek",
    "price_nok",
    "price_eur",
    "barcode",
    "confidence",
    "needs_review",
    "notes",
]


def write_xlsx(rows: Iterable[ProductRow], output_path: str) -> None:
    workbook = Workbook()
    products_sheet = workbook.active
    products_sheet.title = "PRODUCTS"
    review_sheet = workbook.create_sheet("REVIEW")

    products_sheet.append(HEADERS)
    review_headers = [
        "source_file",
        "page",
        "section",
        "product_name_en",
        "variant",
        "designer",
        "art_no",
        "price_eur",
        "needs_review",
        "exported",
        "notes",
        "raw_block_id",
        "raw_snippet",
    ]
    review_sheet.append(review_headers)

    sorted_rows = sorted(rows, key=_sort_key)
    visible_rows = [row for row in sorted_rows if not getattr(row, "noise", False)]
    exported_rows = [row for row in visible_rows if row.exported]
    for row in exported_rows:
        row_dict = row.to_dict()
        products_sheet.append([row_dict.get(header) for header in HEADERS])
    for row in visible_rows:
        if row.needs_review or not row.exported:
            row_dict = row.to_dict()
            review_sheet.append([row_dict.get(header) for header in review_headers])

    workbook.save(output_path)


def write_diagnostic_summary(
    output_path: str,
    source_file: str,
    cached_pages: Iterable,
    cache_meta: dict,
    attempt_results: List[dict],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "SUMMARY"
    summary_sheet.append(["key", "value"])
    summary_sheet.append(["source_file", source_file])
    summary_sheet.append(["num_pages", cache_meta.get("num_pages", 0)])
    summary_sheet.append(["sample_count", cache_meta.get("sample_count", 0)])
    summary_sheet.append(["scan_mode", cache_meta.get("scan_mode", "")])
    summary_sheet.append(
        ["pages_sampled", _format_page_list(cache_meta.get("pages_sampled", []))]
    )
    summary_sheet.append(
        ["top_k_pages", _format_page_list(cache_meta.get("top_k_pages", []))]
    )
    summary_sheet.append(
        ["top_k_scores", _format_score_list(cache_meta.get("top_k_scores", []))]
    )
    summary_sheet.append(
        ["export_policy_mode", cache_meta.get("export_policy_mode", "")]
    )
    summary_sheet.append(["rows_total", cache_meta.get("rows_total", 0)])
    summary_sheet.append(["rows_exported", cache_meta.get("rows_exported", 0)])
    summary_sheet.append(["rows_review", cache_meta.get("rows_review", 0)])
    summary_sheet.append(["rows_noise", cache_meta.get("rows_noise", 0)])
    guardrail_counts = cache_meta.get("guardrail_counts", {}) or {}
    summary_sheet.append(
        ["year_price_blocked", guardrail_counts.get("year_price_blocked", 0)]
    )
    summary_sheet.append(
        ["filtered_color_temp_code", guardrail_counts.get("filtered_color_temp_code", 0)]
    )
    summary_sheet.append(
        ["filtered_short_grade_token", guardrail_counts.get("filtered_short_grade_token", 0)]
    )
    summary_sheet.append(
        ["filtered_watt_code", guardrail_counts.get("filtered_watt_code", 0)]
    )
    summary_sheet.append(
        ["filtered_socket_code", guardrail_counts.get("filtered_socket_code", 0)]
    )
    summary_sheet.append(
        ["filtered_t_series_code", guardrail_counts.get("filtered_t_series_code", 0)]
    )
    summary_sheet.append(
        ["filtered_single_letter_code", guardrail_counts.get("filtered_single_letter_code", 0)]
    )
    summary_sheet.append(
        ["table_stitcher_used_pages", guardrail_counts.get("table_stitcher_used_pages", 0)]
    )
    summary_sheet.append(
        ["table_stitcher_rows_exported", guardrail_counts.get("table_stitcher_rows_exported", 0)]
    )
    summary_sheet.append(
        ["partial_export", guardrail_counts.get("partial_export", 0)]
    )

    pages_sheet = workbook.create_sheet("TOP_K")
    pages_sheet.append(
        [
            "page",
            "signal_score",
            "table_likelihood",
            "numeric_density",
            "price_like_count",
            "cooccurrence_count",
            "mixed_code_count",
            "text_len",
            "table_hint",
            "ocr_used",
        ]
    )
    for page in cached_pages:
        pages_sheet.append(
            [
                getattr(page, "page_number", None),
                getattr(page, "signal_score", 0.0),
                getattr(page, "table_likelihood", 0.0),
                getattr(page, "numeric_density", 0.0),
                getattr(page, "price_like_count", 0),
                getattr(page, "cooccurrence_count", 0),
                getattr(page, "mixed_code_count", 0),
                getattr(page, "text_len", 0),
                bool(getattr(page, "table_hint", False)),
                bool(getattr(page, "ocr_used", False)),
            ]
        )

    attempts_sheet = workbook.create_sheet("ATTEMPTS")
    attempts_sheet.append(
        [
            "parser",
            "ok",
            "reason",
            "eval_score",
            "eval_rows",
            "eval_pages",
            "eval_time_ms",
        ]
    )
    for attempt in attempt_results or []:
        metrics = attempt.get("metrics", {}) or {}
        attempts_sheet.append(
            [
                attempt.get("parser", ""),
                bool(attempt.get("ok")),
                attempt.get("reason", ""),
                metrics.get("eval_score", 0.0),
                metrics.get("eval_rows", 0),
                metrics.get("eval_pages_used", 0),
                metrics.get("eval_time_ms", 0),
            ]
        )

    workbook.save(output_path)


def _format_page_list(pages: Iterable[int], limit: int = 200) -> str:
    items = [str(page) for page in pages if page is not None]
    if len(items) > limit:
        items = items[:limit] + ["..."]
    return ",".join(items)


def _format_score_list(scores: Iterable, limit: int = 200) -> str:
    items = [f"{float(score):.3f}" for score in scores if score is not None]
    if len(items) > limit:
        items = items[:limit] + ["..."]
    return ",".join(items)


def _sort_key(row: ProductRow):
    page = row.page if row.page is not None else 0
    section = (row.section or "").lower()
    art_no = text_utils.canonicalize_art_no(row.art_no)
    return (page, section, art_no)
